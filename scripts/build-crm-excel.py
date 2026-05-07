#!/usr/bin/env python3
"""
build-crm-excel.py
Genera il file CRM Excel da data/zmel-fr-med.json + data/amp-it.json + data/sardegna-prrpt.json
con convenzione naming: crm_seainnova_mappa_boe_YYYY-MM-DD_REVN.xlsx

Uso:
    python3 scripts/build-crm-excel.py [REV]

Se REV non viene fornito, viene auto-incrementato dall'ultima rev trovata in data/.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json
import os
import sys
import re
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / 'data'
STORICO = DATA / 'storico_revisioni'

def detect_next_rev():
    """Trova la rev più alta in data/ + storico_revisioni/ e restituisce la prossima."""
    revs = []
    for d in (DATA, STORICO):
        if not d.exists(): continue
        for f in d.glob('crm_seainnova_mappa_boe_*_REV*.xlsx'):
            m = re.search(r'_REV(\d+)\.xlsx$', f.name)
            if m: revs.append(int(m.group(1)))
    return max(revs) + 1 if revs else 1

def archive_current_to_storico():
    """Sposta il file attuale (se esiste) in storico_revisioni/."""
    STORICO.mkdir(exist_ok=True)
    for f in DATA.glob('crm_seainnova_mappa_boe_*_REV*.xlsx'):
        target = STORICO / f.name
        if not target.exists():
            f.rename(target)
            print(f'  Archiviato: {f.name} → storico_revisioni/')

def build():
    rev = int(sys.argv[1]) if len(sys.argv) > 1 else detect_next_rev()
    today = date.today().isoformat()
    filename = f'crm_seainnova_mappa_boe_{today}_REV{rev}.xlsx'
    out_path = DATA / filename

    zmel = json.loads((DATA / 'zmel-fr-med.json').read_text())
    amp  = json.loads((DATA / 'amp-it.json').read_text())
    sard = json.loads((DATA / 'sardegna-prrpt.json').read_text())

    rows = []; n = 0

    def products(cat):
        return {
            'amp_cliente':'Betty Buoys, Sentinel',
            'in_progress':'Betty Buoys (sistema completo nuovo impianto)',
            'private':'Betty Buoys, Watch Buoy',
            'amp_pnrr':'Betty Buoys (espansione PNRR)',
        }.get(cat,'Betty Buoys, Sentinel')

    def score(entry, has_email, has_phone):
        cap = entry.get('capacity', {})
        total = cap.get('total') or 0
        notes = (entry.get('notes') or '').lower()
        is_client = 'cliente' in notes or 'crm' in notes or 'installazione betty' in notes
        is_pnrr = 'pnrr' in notes
        is_progress = entry.get('status') == 'in-progress' or 'progett' in notes
        has_contact = has_email or has_phone
        if is_client: return 10
        if is_pnrr: return 9
        if is_progress: return 9
        if total >= 100: return 9
        if total >= 50: return 8
        if has_contact and total >= 20: return 8
        if has_contact: return 7
        if total >= 20: return 7
        return 6

    def add(idx, ente, ref, ruolo, em, tel, prod, zona, sc, st, pri, fonte, ogg, note):
        return (idx, ente, ref, ruolo, em, tel, prod, zona, sc, st, pri, None, fonte, 'nuovo', ogg, None, None, note)

    def parse_role(referente):
        if not referente: return None
        if '—' in referente:
            parts = referente.split('—')
            if len(parts) >= 2:
                r = parts[1].strip().split('.')[0].split('(')[0].strip()
                return r if len(r) < 80 else None
        return None

    def parse_name(referente):
        if not referente: return None
        for sep in ['—','. Presidente',', Presidente','. Risorse','. Pres','. Operatori','. Commissario','. Subentrato','. Capitainerie',';']:
            if sep in referente:
                return referente.split(sep)[0].strip()
        return referente

    # ZMEL FR
    for z in zmel['buoyFields']:
        n += 1
        m = z.get('manager') or {}
        has_email = bool(m.get('email')); has_phone = bool(m.get('phone'))
        is_progress = z.get('status') == 'in-progress'
        cap = z.get('capacity', {})
        total = cap.get('total')
        cap_str = f"{total} boe" if total else "in progettazione"
        note_parts = [f"Mappa Boe app — {cap_str}"]
        if cap.get('breakdown'): note_parts.append(cap['breakdown'][:120])
        if z.get('notes'): note_parts.append(z['notes'][:200])
        if m.get('address'): note_parts.append(f"Sede: {m['address']}")
        if m.get('vhf'): note_parts.append(f"VHF: {m['vhf']}")
        if m.get('phone_mobile'): note_parts.append(f"Mobile: {m['phone_mobile']}")
        if m.get('phone_emergency'): note_parts.append(f"Urgenza: {m['phone_emergency']}")
        if m.get('phone_op'): note_parts.append(f"Tel op: {m['phone_op']}")
        if m.get('email_alt'): note_parts.append(f"Email alt: {m['email_alt']}")
        if m.get('email_op'): note_parts.append(f"Email op: {m['email_op']}")
        sc = score(z, has_email, has_phone)
        cat = 'in_progress' if is_progress else 'zmel_op'
        ref_full = m.get('contact_person')
        rows.append(add(n, z['name'], parse_name(ref_full), parse_role(ref_full),
            m.get('email'), m.get('phone'), products(cat),
            f"{z.get('dept_name','')}, Francia", sc, 'da_contattare',
            'alta' if sc >= 9 else ('media' if sc >= 7 else 'bassa'),
            'mappa_boe_app',
            f"Betty Buoys — {'progetto' if is_progress else 'campo boe'} {z['name']}",
            ' | '.join(note_parts)))

    # AMP IT
    for a in amp['ampFields']:
        n += 1
        m = a.get('manager') or {}
        has_email = bool(m.get('email')); has_phone = bool(m.get('phone'))
        notes = (a.get('notes') or '').lower()
        is_client = 'cliente' in notes or 'installazione betty' in notes
        is_pnrr = 'pnrr' in notes
        is_progress = a.get('status') == 'in-progress'
        is_private = a.get('subtype') == 'private'
        cap = a.get('capacity', {})
        total = cap.get('total')
        cap_str = f"{total} boe" if total else "n/d"
        note_parts = [f"Mappa Boe app — {cap_str}"]
        if cap.get('breakdown'): note_parts.append(cap['breakdown'][:120])
        if a.get('notes'): note_parts.append(a['notes'][:200])
        if m.get('address'): note_parts.append(f"Sede: {m['address']}")
        if m.get('pec'): note_parts.append(f"PEC: {m['pec']}")
        if m.get('fax'): note_parts.append(f"Fax: {m['fax']}")
        if m.get('email_alt'): note_parts.append(f"Email alt: {m['email_alt']}")
        if m.get('phone_alt'): note_parts.append(f"Tel alt: {m['phone_alt']}")
        if m.get('phone_mobile'): note_parts.append(f"Mobile: {m['phone_mobile']}")
        sc = score(a, has_email, has_phone)
        if is_client: cat='amp_cliente'
        elif is_pnrr: cat='amp_pnrr'
        elif is_progress: cat='in_progress'
        elif is_private: cat='private'
        else: cat='amp_op'
        ref_full = m.get('contact_person')
        rows.append(add(n, a['name'], parse_name(ref_full), parse_role(ref_full),
            m.get('email'), m.get('phone'), products(cat),
            f"{a.get('region','')}, Italia", sc, 'da_contattare',
            'alta' if sc >= 9 else ('media' if sc >= 7 else 'bassa'),
            'mappa_boe_app',
            f"Betty Buoys — {'progetto ' if is_progress else ''}AMP {a['name']}",
            ' | '.join(note_parts)))

    # Sardegna PRRPT
    for s in sard['buoyFields']:
        n += 1
        m = s.get('manager') or {}
        has_email = bool(m.get('email')); has_phone = bool(m.get('phone'))
        cap = s.get('capacity', {})
        total = cap.get('total') or 0
        note_parts = [f"Mappa Boe app — PRRPT 2024 Sardegna: {total} posti"]
        note_parts.append(f"Comune {s.get('comune','')} ({s.get('quadrant','')})")
        if s.get('notes'): note_parts.append(s['notes'][:200])
        if m.get('address'): note_parts.append(f"Sede: {m['address']}")
        if m.get('phone_porto_istana'): note_parts.append(f"Tel Porto Istana: {m['phone_porto_istana']}")
        if m.get('phone_diving'): note_parts.append(f"Tel Diving: {m['phone_diving']}")
        sc = 9 if total >= 80 else (8 if total >= 30 else 7 if total >= 15 else 6)
        ref_full = m.get('contact_person')
        rows.append(add(n, f"Campo Boe {s['name']}", parse_name(ref_full), parse_role(ref_full),
            m.get('email'), m.get('phone'), 'Betty Buoys, Watch Buoy',
            f"{s.get('comune','')}, Sardegna", sc, 'da_contattare',
            'alta' if sc >= 9 else ('media' if sc >= 7 else 'bassa'),
            'mappa_boe_app_PRRPT2024',
            f"Betty Buoys — campo boe {s['name']} ({s.get('comune','')})",
            ' | '.join(note_parts)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mappa Boe'
    headers = ['#', 'Ente', 'Referente', 'Ruolo', 'Email', 'Telefono', 'Prodotti', 'Zona',
               'Score', 'Stato', 'Priorita', 'Data Contatto', 'Fonte', 'Stato Bozza',
               'Oggetto Email', 'Data Bozza', 'Data Invio', 'Note Followup']
    ws.append(headers)
    header_fill = PatternFill(start_color='03213F', end_color='03213F', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for col in range(1, len(headers)+1):
        c = ws.cell(row=1, column=col); c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal='left', vertical='center')
    for r in rows: ws.append(r)
    widths = [4, 42, 32, 18, 42, 22, 28, 26, 7, 16, 12, 12, 22, 14, 50, 12, 12, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    score_fills = {
        10: PatternFill(start_color='1A7BB8', end_color='1A7BB8', fill_type='solid'),
        9:  PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid'),
        8:  PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        7:  PatternFill(start_color='E8F4FA', end_color='E8F4FA', fill_type='solid'),
    }
    for r_idx in range(2, len(rows)+2):
        sc_cell = ws.cell(row=r_idx, column=9)
        if sc_cell.value in score_fills:
            sc_cell.fill = score_fills[sc_cell.value]
            if sc_cell.value in (10, 9):
                sc_cell.font = Font(bold=True, color='FFFFFF')
    ws.freeze_panes = 'A2'

    archive_current_to_storico()
    wb.save(out_path)

    emails = sum(1 for r in rows if r[4])
    phones = sum(1 for r in rows if r[5])
    refs = sum(1 for r in rows if r[2])
    ruoli = sum(1 for r in rows if r[3])
    print(f'\n✔ {out_path.name}')
    print(f'  {n} contatti | Email: {emails}/{n} ({emails*100//n}%) | Tel: {phones}/{n} ({phones*100//n}%) | Ref: {refs}/{n} ({refs*100//n}%) | Ruolo: {ruoli}/{n} ({ruoli*100//n}%)')

if __name__ == '__main__':
    build()
